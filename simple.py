import openpyxl

# 打开文件
wb = openpyxl.load_workbook(filename="testpyhon.xlsx")
# 获取活动的sheet页
active_ws = wb.active

# 循环开始按行读取表格,把表格按行存到一个列表中,列表中的每个项目又是一个列表，子列表是表格的行数据
datalst = []
for row in active_ws.rows:
    tmplst = []
    for cell in row:
        if cell.value != None:
            tmplst.append(cell.value)
    datalst.append(tmplst)

# 开始构造要输出的数据,从第二行开始,并写到一个以保单号作为key的字典里面
datadict = dict()
for item in datalst[1:]:  # 不要第一个项目“标题”
    if item[1] in datadict:  # 用保单号作为索引
        datadict[item[1]] = datadict[item[1]] + item[2:]  # 如果保单号在字典里，就将后6列数据补充到已有字典数据的后面
    else:
        datadict[item[1]] = item[2:]  # 如果保单号不在字典里，就将后6列数据赋给索引

###########################
outlst = []  # 构造最终输出的数据
idx = 1  # idx是序号
tmplst = []  # 清空临时列表
for k, v in datadict.items():  # 将字典中的数据加上编号存储到outlst中
    tmplst = [idx, k] + v
    outlst.append(tmplst)
    idx = idx + 1

# 开始输出，写到一个新文件里面去
outputwb = openpyxl.Workbook()
outputsheet = outputwb.active
for row in range(len(outlst)):
    for col in range(len(outlst[row])):
        outputsheet.cell(row=row + 1, column=col + 1).value = outlst[row][col]
outputwb.save(r'output.xlsx')
print('数据输出完成')