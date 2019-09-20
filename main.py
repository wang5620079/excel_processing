# coding=utf-8
# -*- coding: utf-8 -*-

import datetime
import openpyxl



def main():
    #打开文件
    wb=openpyxl.load_workbook(filename="testpyhon.xlsx")
    #获取活动的sheet页
    act_sheet=wb.active

    ##########################
    print('###########开始读取数据##########')
    #循环开始按行读取数据,把所有数据存到一个list中,list中的数据是一行数据的list
    datalst=[]
    for row in act_sheet.rows:
        tmplst=[]
        for cell in row:
            if None!=cell.value and len(str(cell.value).strip())>0:
                tmplst.append(cell.value)
            else:
                tmplst.append('')
        datalst.append(tmplst)
    print('###########读取数据完成##########')
    # for item in datalst:
    #     print(item)

    ##########################
    print('###########开始构造输出数据##########')
    #开始构造要输出的数据,从第二行开始,并写到一个以保单号作为key的字典里面
    #datadict是我们要获取的中间结果
    datadict=dict()
    tmpdict=dict()
    for item in datalst[1:]:
        #如果字典里已经有了，则扩展，如果没有，就加进去,数据格式为{保单号:[[第一个6列数据]，[第二个6列数据]，...]}
        if item[1] in tmpdict:
            #为了读起来方便，在这里进行了一次排序
            tmpdict[item[1]].append(item[2:])
        else:
            tmpdict[item[1]]=[item[2:],]
    for k, v in tmpdict.items():
        # 对每一个保单，根据次序进行排序
        v = v.sort(key=lambda x: x[0])
    for k,v in tmpdict.items():
        tmplst=[]
        for item in v:
            tmplst=tmplst+item
        datadict[k]=tmplst

    ##########################
    #构造新数据的标题头
    titlelst=datalst[0]
    expendpart=titlelst[2:]
    #找到数据中，列数最大的列的个数（除去序号，保单号）
    # maxcol=max([len(item) for item in datadict.values()])
    maxcol=max([len(item) for item in datadict.values()])
    #在这里找到最大的列数是156列，即26个“次序 缴费期数 代码 日期 币种 金额”，构造26个“次序 缴费期数 代码 日期 币种 金额”，它的格式是“次序* 缴费期数* 代码* 日期* 币种* 金额*”，*是数字
    expendpartlst=[]
    for i in range(int(maxcol/len(expendpart))):
        for item in expendpart:
            expendpartlst.append(item+str(i+1))


    ###########################
    #构造最终输出的数据
    #第一步，先写上标题
    outlst=[]
    #第二步，填写数据
    #序号初始都是0
    for k, v in datadict.items():
        tmplst = [0, k] + v
        outlst.append(tmplst)
    #为了便于阅读，根据保单号进行排序
    outlst.sort(key=lambda x:x[1])
    #改写序号
    for i in range(len(outlst)):
        outlst[i][0]=i
    #第三步：写上标题,把标题作为第一个插入进去
    outlst.insert(0,titlelst[0:2] + expendpartlst)
    print('###########构造输出数据完成##########')
    for item in outlst:
        print(item)
    print('###########开始写入数据##########')
    #开始输出，写到一个新文件里面去
    outwb=openpyxl.Workbook()
    outsheet=outwb.active
    for row in range(len(outlst)):
        for col  in range(len(outlst[row])):
            if isinstance(outlst[row][col],datetime.datetime):
                outsheet.cell(row=row + 1, column=col + 1).value=outlst[row][col].strftime('%Y/%m/%d')
            else:
                outsheet.cell(row=row+1,column=col+1).value=outlst[row][col]
    outwb.save(r'output.xlsx')
    print('###########写入数据完成##########')

if __name__=="__main__":
    main()